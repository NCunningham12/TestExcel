from openpyxl import Workbook, load_workbook

wb = load_workbook('STO_BOB.xlsx')
ws = wb.active

ws.title = "STO LLC Auction Calculations"

truckCategories = [["ID", "Auction Price", "Province", "Canadian Price (CAD)", "US Price (USD)", "Book (USD)", "Total US (USD)", "BOB"]]

# for row in truckCategories:
#   ws.append(row)

# INPUTS: Auction Price, Province, US Price, Book
# OUTPUTS: ID, Canadian Price, Total US, BOB

def BOB(): 
  id = 0
  
  for row in range(2, ws.max_row + 1):
    id += 1
    ws['A' + str(row)].value = id                                     # Creates a unique ID number for each vehicle for reference only

    auction_price = ws['B' + str(row)].value
    province = ws['C' + str(row)].value
    us_price = ws['G' + str(row)].value
    book_cad = ws['D' + str(row)].value
    exchange = ws['K2'].value

    book_usd = book_cad * exchange
    ws['E' + str(row)].value = book_usd


    if province.lower() == 'alb' or province.lower() == 'bc':         # Checks if province is WEST
      auction_price += 1000
      auction_price *= 1.05                                           # Adds 1000 and then adds 5% of the new total
      canadian_price = round(auction_price, 2)
      ws['F' + str(row)] = round(canadian_price, 2)

      us_price = round(canadian_price * exchange, 2)                  # Converts the CAD to USD based on the input "Exchange rate"
      ws['G' + str(row)] = round(us_price, 2)

      us_price += 2300                                                
      us_price *= 1.01                                                # Adds 2300 and then adds 1% of the new total
      total_us_price = round(us_price, 2)  
      ws['H' + str(row)] = total_us_price                       
    elif province.lower() == 'ont':                                   # Checks if province is EAST
      auction_price += 1000
      auction_price *= 1.13                                           # Adds 1000 and then adds 13% of the new total
      canadian_price = round(auction_price, 2)                        
      ws['F' + str(row)] = round(canadian_price, 2)

      us_price = round(canadian_price * exchange)                     # Converts the CAD to USD based on the input "Exchange Rate"
      ws['G' + str(row)] = round(us_price, 2)

      us_price += 3400                                                
      us_price *= 1.01                                                # Adds 3400 and then adds 1% of the new total
      total_us_price = round(us_price, 2)
      ws['H' + str(row)] = total_us_price

    bob = round(book_usd - total_us_price, 2)                             # Subtracts Total US from Book to calculate BOB
    ws['I' + str(row)] = bob


    # print("Auction Price: " + str(round(auction_price, 2)))
    # print("Canadian Price: " + str(canadian_price))
    # print("US Price: " + str(round(us_price, 2)))
    # print("Total US Price: " + str(total_us_price))
    # print("BOB: " + str(bob))

BOB()

wb.save('STO_BOB.xlsx')
print("✅ WORKBOOK UPDATED")